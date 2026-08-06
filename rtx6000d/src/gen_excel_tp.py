#!/usr/bin/env python3
"""TP1/TP2/TP4 comparison + communication analysis -> append sheets to Excel."""
import json, gzip, re
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = "/home/tina/DEV/microbench/results"
CFG = {
  ("tp1",1): {
    ("prefill",1): (f"{ROOT}/dsv4_prof/dp0_pp0_tp0_dcp0_ep0_rank0.1786033965325236324.pt.trace.json.gz",1,16,55.0),
    ("decode",1):  (f"{ROOT}/dsv4_prof/dp0_pp0_tp0_dcp0_ep0_rank0.1786033983186682170.pt.trace.json.gz",16,16,73.0),
    ("prefill",64):(f"{ROOT}/dsv4_prof_b64/dp0_pp0_tp0_dcp0_ep0_rank0.1786037877388495510.pt.trace.json.gz",1,1024,360.0),
    ("decode",64): (f"{ROOT}/dsv4_prof_b64/dp0_pp0_tp0_dcp0_ep0_rank0.1786037957570883410.pt.trace.json.gz",16,1024,220.0),
  },
  ("tp2",2): {
    ("prefill",1): (f"{ROOT}/dsv4_prof_tp2/dp0_pp0_tp0_dcp0_ep0_rank0.1786039282820340968.pt.trace.json.gz",1,16,60.0),
    ("decode",1):  (f"{ROOT}/dsv4_prof_tp2/dp0_pp0_tp0_dcp0_ep0_rank0.1786039305795217019.pt.trace.json.gz",16,16,80.0),
    ("prefill",64):(f"{ROOT}/dsv4_prof_tp2/dp0_pp0_tp0_dcp0_ep0_rank0.1786039340401701700.pt.trace.json.gz",1,1024,256.0),
    ("decode",64): (f"{ROOT}/dsv4_prof_tp2/dp0_pp0_tp0_dcp0_ep0_rank0.1786039474598200593.pt.trace.json.gz",16,1024,182.0),
  },
  ("tp4",4): {
    ("prefill",1): (f"{ROOT}/dsv4_prof_tp4/dp0_pp0_tp0_dcp0_ep0_rank0.1786040196278428996.pt.trace.json.gz",1,16,59.0),
    ("decode",1):  (f"{ROOT}/dsv4_prof_tp4/dp0_pp0_tp0_dcp0_ep0_rank0.1786040224303881676.pt.trace.json.gz",16,16,82.0),
    ("prefill",64):(f"{ROOT}/dsv4_prof_tp4/dp0_pp0_tp0_dcp0_ep0_rank0.1786040264965651555.pt.trace.json.gz",1,1024,255.0),
    ("decode",64): (f"{ROOT}/dsv4_prof_tp4/dp0_pp0_tp0_dcp0_ep0_rank0.1786040402221117991.pt.trace.json.gz",16,1024,207.0),
  },
}

def load(p):
    with gzip.open(p,"rt") as f: return json.load(f)["traceEvents"]

def kind(n):
    if "ncclDevKernel_AllReduce" in n: return "comm_nccl_allreduce"
    if "ncclDevKernel_AllGather" in n: return "comm_nccl_allgather"
    if "ncclDevKernel" in n: return "comm_nccl_other"
    if "pcie_allreduce" in n: return "comm_pcie_ar"
    if "moefusedsilu" in n: return "moe_nvfp4"
    if "gemmdense" in n: return "fp8_dense"
    if "attentionmla" in n: return "attn_mla"
    if "gemvx" in n or "cutlass_80_tensorop_bf16" in n: return "lm_head"
    if "gumbel" in n or "sample" in n: return "sampler"
    return "other"

data = {}
for (label,tp),scenes in CFG.items():
    for (sc,bs),(path,steps,toks,wall) in scenes.items():
        ev=[e for e in load(path) if e.get("ph")=="X" and e.get("cat") in ("kernel","gpu_memcpy","gpu_memset")]
        busy=sum(e["dur"] for e in ev)
        agg=defaultdict(lambda:[0.0,0])
        for e in ev:
            k=kind(e["name"]); agg[k][0]+=e["dur"]; agg[k][1]+=1
        data[(label,sc,bs)]={"tp":tp,"busy":busy,"wall":wall,"toks":toks,"steps":steps,"agg":agg}

bold=Font(bold=True); hdr=PatternFill("solid",fgColor="D9E1F2"); comm_fill=PatternFill("solid",fgColor="F8CBAD")
wb=load_workbook(f"{ROOT}/DeepSeekV4_NVFP4_profiling_comparison.xlsx")
for name in ["TP扩展总览","TP通信分析"]:
    if name in wb.sheetnames: del wb[name]

# ---- TP overview ----
ws=wb.create_sheet("TP扩展总览")
ws.append(["场景","batch","TP","GPU busy (ms)","vs TP1 加速比","wall (ms)","wall vs TP1","吞吐 tok/s (wall)","通信 (ms)","通信占 busy","MoE NVFP4 (ms)","MoE 占 busy"])
for sc in ["prefill","decode"]:
    for bs in [1,64]:
        base=data[("tp1",sc,bs)]
        for label in ["tp1","tp2","tp4"]:
            d=data[(label,sc,bs)]
            comm=sum(v[0] for k,v in d["agg"].items() if k.startswith("comm_"))
            moe=d["agg"].get("moe_nvfp4",[0,0])[0]
            ws.append([sc,bs,d["tp"],round(d["busy"]/1e3,2),round(base["busy"]/d["busy"],2),
                       d["wall"],round(base["wall"]/d["wall"],2),round(d["toks"]/d["wall"]*1e3),
                       round(comm/1e3,2),f"{100*comm/d['busy']:.1f}%",
                       round(moe/1e3,2),f"{100*moe/d['busy']:.1f}%"])
            if comm>0:
                for cell in ws[ws.max_row][8:10]: cell.fill=comm_fill
for c in ws[1]:
    if c.value: c.font=bold; c.fill=hdr

# ---- comm analysis ----
ws=wb.create_sheet("TP通信分析")
ws.append(["场景","batch","TP","NCCL AllReduce ms(calls)","NCCL AllGather ms(calls)","b12x PCIe-AR ms(calls)","通信合计 ms","通信占 busy","单次 AllReduce 均值 us","说明"])
notes={
 ("decode",64,2):"AllReduce 为主;MoE expert 按 TP 切半",
 ("decode",64,4):"NCCL AR 22.4ms + PCIe-AR 14.3ms 双路径;PCIe ring 是瓶颈",
 ("prefill",64,4):"prefill 每步 AR payload 大(1024 tok),54% busy 是通信",
 ("decode",1,4):"小 payload 走 b12x pcie_allreduce(80 次,~15.6us/次)",
}
for sc in ["prefill","decode"]:
    for bs in [1,64]:
        for label in ["tp1","tp2","tp4"]:
            d=data[(label,sc,bs)]
            ar=d["agg"].get("comm_nccl_allreduce",[0,0]); ag=d["agg"].get("comm_nccl_allgather",[0,0]); pa=d["agg"].get("comm_pcie_ar",[0,0])
            comm=ar[0]+ag[0]+pa[0]+d["agg"].get("comm_nccl_other",[0,0])[0]
            ws.append([sc,bs,d["tp"],
                       f"{ar[0]/1e3:.2f}({ar[1]})",f"{ag[0]/1e3:.2f}({ag[1]})",f"{pa[0]/1e3:.2f}({pa[1]})",
                       round(comm/1e3,2),f"{100*comm/d['busy']:.1f}%",
                       round(ar[0]/ar[1],1) if ar[1] else "-",
                       notes.get((sc,bs,d["tp"]),"")])
            if comm>0:
                for cell in ws[ws.max_row][7:9]: cell.fill=comm_fill
for c in ws[1]:
    if c.value: c.font=bold; c.fill=hdr

for ws in [wb["TP扩展总览"],wb["TP通信分析"]]:
    for i,col in enumerate(ws.columns,1):
        w=max((len(str(c.value)) for c in col if c.value is not None),default=8)
        ws.column_dimensions[get_column_letter(i)].width=min(w+2,50)

wb.save(f"{ROOT}/DeepSeekV4_NVFP4_profiling_comparison.xlsx")
print("Excel updated")

# text dump
print(f"{'scene':<8}{'bs':<4}{'tp':<3}{'busy_ms':>8}{'speedup':>8}{'wall':>6}{'comm_ms':>8}{'comm%':>7}{'moe_ms':>8}{'moe%':>7}")
for sc in ["prefill","decode"]:
    for bs in [1,64]:
        base=data[("tp1",sc,bs)]
        for label in ["tp1","tp2","tp4"]:
            d=data[(label,sc,bs)]
            comm=sum(v[0] for k,v in d["agg"].items() if k.startswith("comm_"))
            moe=d["agg"].get("moe_nvfp4",[0,0])[0]
            print(f"{sc:<8}{bs:<4}{d['tp']:<3}{d['busy']/1e3:>8.2f}{base['busy']/d['busy']:>8.2f}{d['wall']:>6.0f}{comm/1e3:>8.2f}{100*comm/d['busy']:>6.1f}%{moe/1e3:>8.2f}{100*moe/d['busy']:>6.1f}%")
