#!/usr/bin/env python3
"""EP (TP2+expert-parallel, marlin backend) vs TP2 (b12x fused MoE) comparison -> append sheet to Excel."""
import json, gzip
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = "/home/tina/DEV/microbench/results"
CFG = {
  ("tp2", "TP2 (b12x MoE)"): {
    ("prefill",1): (f"{ROOT}/dsv4_prof_tp2/dp0_pp0_tp0_dcp0_ep0_rank0.1786039282820340968.pt.trace.json.gz",1,16,60.0),
    ("decode",1):  (f"{ROOT}/dsv4_prof_tp2/dp0_pp0_tp0_dcp0_ep0_rank0.1786039305795217019.pt.trace.json.gz",16,16,80.0),
    ("prefill",64):(f"{ROOT}/dsv4_prof_tp2/dp0_pp0_tp0_dcp0_ep0_rank0.1786039340401701700.pt.trace.json.gz",1,1024,256.0),
    ("decode",64): (f"{ROOT}/dsv4_prof_tp2/dp0_pp0_tp0_dcp0_ep0_rank0.1786039474598200593.pt.trace.json.gz",16,1024,182.0),
  },
  ("tp2ep", "TP2+EP (marlin MoE)"): {
    ("prefill",1): (f"{ROOT}/dsv4_prof_tp2ep/dp0_pp0_tp0_dcp0_ep0_rank0.1786044459848868665.pt.trace.json.gz",1,16,48.0),
    ("decode",1):  (f"{ROOT}/dsv4_prof_tp2ep/dp0_pp0_tp0_dcp0_ep0_rank0.1786044480929469583.pt.trace.json.gz",16,16,77.0),
    ("prefill",64):(f"{ROOT}/dsv4_prof_tp2ep/dp0_pp0_tp0_dcp0_ep0_rank0.1786044504362448383.pt.trace.json.gz",1,1024,240.0),
    ("decode",64): (f"{ROOT}/dsv4_prof_tp2ep/dp0_pp0_tp0_dcp0_ep0_rank0.1786044614320442436.pt.trace.json.gz",16,1024,175.0),
  },
}

def load(p):
    with gzip.open(p,"rt") as f: return json.load(f)["traceEvents"]

def kind(n):
    if "ncclDevKernel_AllReduce" in n: return "comm_nccl_allreduce"
    if "ncclDevKernel_AllGather" in n: return "comm_nccl_allgather"
    if "ncclDevKernel_ReduceScatter" in n: return "comm_nccl_reducescatter"
    if "ncclDevKernel_AllToAll" in n: return "comm_nccl_all2all"
    if "ncclDevKernel" in n: return "comm_nccl_other"
    if "pcie_allreduce" in n: return "comm_pcie_ar"
    if "marlin_moe" in n or "Marlin" in n: return "moe_marlin"
    if "moefusedsilu" in n: return "moe_nvfp4"
    if "moe_align_block_size" in n or "topkGating" in n: return "moe_route"
    if "gemmdense" in n: return "fp8_dense"
    if "attentionmla" in n: return "attn_mla"
    if "gemvx" in n or "cutlass_80_tensorop_bf16" in n or "cutlass_80_wmma" in n: return "bf16_gemm"
    if "gumbel" in n or "sample" in n: return "sampler"
    return "other"

data = {}
for (label, desc), scenes in CFG.items():
    for (sc,bs),(path,steps,toks,wall) in scenes.items():
        ev=[e for e in load(path) if e.get("ph")=="X" and e.get("cat") in ("kernel","gpu_memcpy","gpu_memset")]
        busy=sum(e["dur"] for e in ev)
        agg=defaultdict(lambda:[0.0,0])
        for e in ev:
            k=kind(e["name"]); agg[k][0]+=e["dur"]; agg[k][1]+=1
        data[(label,sc,bs)]={"busy":busy,"wall":wall,"toks":toks,"steps":steps,"agg":agg}

bold=Font(bold=True); hdr=PatternFill("solid",fgColor="D9E1F2"); comm_fill=PatternFill("solid",fgColor="F8CBAD")
wb=load_workbook(f"{ROOT}/DeepSeekV4_NVFP4_profiling_comparison.xlsx")
for name in ["EP对比"]:
    if name in wb.sheetnames: del wb[name]

ws=wb.create_sheet("EP对比")
ws.append(["场景","batch","配置","GPU busy (ms)","wall (ms)","通信 (ms)","通信占 busy",
           "MoE (ms)","MoE kernel","AR ms(calls)","AG ms(calls)","PCIe-AR ms(calls)","说明"])
notes={
 ("prefill",1,"tp2ep"):"marlin W4A16 MoE;小流量通信少",
 ("decode",1,"tp2ep"): "PCIe-AR 均 494us/次(vs TP2 15.6us)——EP 合并全量 hidden 输出,AR payload 不随 rank 缩小",
 ("prefill",64,"tp2ep"):"NCCL AR 均 4.59ms/次,通信 82% busy;dispatch 走 AllGather(非 AllToAll)",
 ("decode",64,"tp2ep"):"marlin MoE 42.4ms 与 b12x 45.4ms 相当,但通信 65.2ms(45.5%)远高于 TP2 的 8.3%",
}
for sc in ["prefill","decode"]:
    for bs in [1,64]:
        for label,desc in [("tp2","TP2 (b12x MoE)"),("tp2ep","TP2+EP (marlin MoE)")]:
            d=data[(label,sc,bs)]
            comm=sum(v[0] for k,v in d["agg"].items() if k.startswith("comm_"))
            moe=d["agg"].get("moe_nvfp4",[0,0])[0]+d["agg"].get("moe_marlin",[0,0])[0]
            ar=d["agg"].get("comm_nccl_allreduce",[0,0]); ag=d["agg"].get("comm_nccl_allgather",[0,0]); pa=d["agg"].get("comm_pcie_ar",[0,0])
            mk = "b12x fused NVFP4" if label=="tp2" else "marlin W4A16 (dequant+bf16)"
            ws.append([sc,bs,desc,round(d["busy"]/1e3,2),d["wall"],round(comm/1e3,2),f"{100*comm/d['busy']:.1f}%",
                       round(moe/1e3,2),mk,
                       f"{ar[0]/1e3:.2f}({ar[1]})",f"{ag[0]/1e3:.2f}({ag[1]})",f"{pa[0]/1e3:.2f}({pa[1]})",
                       notes.get((sc,bs,label),"")])
            if comm>0:
                for cell in ws[ws.max_row][5:7]: cell.fill=comm_fill
for c in ws[1]:
    if c.value: c.font=bold; c.fill=hdr
for i,col in enumerate(ws.columns,1):
    w=max((len(str(c.value)) for c in col if c.value is not None),default=8)
    ws.column_dimensions[get_column_letter(i)].width=min(w+2,60)
wb.save(f"{ROOT}/DeepSeekV4_NVFP4_profiling_comparison.xlsx")
print("Excel updated")

# text dump
print(f"{'scene':<8}{'bs':<4}{'cfg':<7}{'busy_ms':>9}{'wall':>6}{'comm_ms':>9}{'comm%':>7}{'moe_ms':>8}")
for sc in ["prefill","decode"]:
    for bs in [1,64]:
        for label in ["tp2","tp2ep"]:
            d=data[(label,sc,bs)]
            comm=sum(v[0] for k,v in d["agg"].items() if k.startswith("comm_"))
            moe=d["agg"].get("moe_nvfp4",[0,0])[0]+d["agg"].get("moe_marlin",[0,0])[0]
            print(f"{sc:<8}{bs:<4}{label:<7}{d['busy']/1e3:>9.2f}{d['wall']:>6.0f}{comm/1e3:>9.2f}{100*comm/d['busy']:>6.1f}%{moe/1e3:>8.2f}")
