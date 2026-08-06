#!/usr/bin/env python3
"""Aggregate b1/b64 prefill/decode traces -> comparison Excel."""
import json, gzip, re, sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

B1 = "/home/tina/DEV/microbench/results/dsv4_prof"
B64 = "/home/tina/DEV/microbench/results/dsv4_prof_b64"
# (path, steps, prefill_tokens); gen_toks = total tokens generated/processed for per-token metrics
TRACES = {
    ("prefill", 1):  (f"{B1}/dp0_pp0_tp0_dcp0_ep0_rank0.1786033965325236324.pt.trace.json.gz", 1, 16),
    ("decode", 1):   (f"{B1}/dp0_pp0_tp0_dcp0_ep0_rank0.1786033983186682170.pt.trace.json.gz", 16, 16),
    ("prefill", 64): (f"{B64}/dp0_pp0_tp0_dcp0_ep0_rank0.1786037877388495510.pt.trace.json.gz", 1, 1024),
    ("decode", 64):  (f"{B64}/dp0_pp0_tp0_dcp0_ep0_rank0.1786037957570883410.pt.trace.json.gz", 16, 1024),
}
WALL = {("prefill",1): 55.0, ("decode",1): 73.0, ("prefill",64): 360.0, ("decode",64): 220.0}  # ms

def load(p):
    with gzip.open(p, "rt") as f: return json.load(f)["traceEvents"]

def classify(n, avg):
    if "moefusedsilu" in n: return "MoE NVFP4 (fused SiLU)"
    if "gemvx" in n: return "lm_head (bf16)" if avg > 100 else "router/gate"
    if "cutlass_80_tensorop_bf16" in n: return "lm_head (bf16)"
    if "cutlass_80_wmma_tensorop_bf16" in n: return "router/gate"
    if "gemmdense" in n: return "attn proj + shared expert (FP8)"
    if "attentionmla" in n: return "MLA attention core"
    if "integrationresidual" in n or "hc_head" in n: return "MHC hyperconnection"
    if "quantize" in n or "QNormRope" in n: return "act quant + inv-RoPE"
    if "topkGating" in n: return "router/gate"
    if "gumbel" in n or "sample" in n: return "sampler"
    if "indexSelect" in n or "index_elementwise" in n: return "embed/gather"
    if "Memcpy" in n or "memcpy" in n or "Fill" in n or "apply_write" in n or "gather_block" in n or "post_update" in n:
        return "memcpy/fill/misc"
    if "elementwise" in n or "direct_copy" in n: return "elementwise"
    return "other"

def aggregate(path, steps):
    ev = [e for e in load(path) if e.get("ph") == "X" and e.get("cat") in ("kernel","gpu_memcpy","gpu_memset")]
    tot = sum(e["dur"] for e in ev)
    t0 = min(e["ts"] for e in ev); t1 = max(e["ts"]+e["dur"] for e in ev)
    by_name = defaultdict(lambda: [0.0,0])
    for e in ev: by_name[e["name"]][0]+=e["dur"]; by_name[e["name"]][1]+=1
    by_cat = defaultdict(lambda: [0.0,0])
    for n,(d,c) in by_name.items():
        by_cat[classify(n, d/c)][0]+=d; by_cat[classify(n, d/c)][1]+=c
    return {"busy_us": tot, "span_us": t1-t0, "cat": by_cat, "kernels": len(ev)}

data = {}
for key, (path, steps, toks) in TRACES.items():
    r = aggregate(path, steps)
    r["steps"]=steps; r["toks"]=toks; r["wall_ms"]=WALL[key]
    data[key] = r

cats = ["MoE NVFP4 (fused SiLU)", "lm_head (bf16)", "attn proj + shared expert (FP8)",
        "MLA attention core", "MHC hyperconnection", "act quant + inv-RoPE",
        "router/gate", "sampler", "embed/gather", "elementwise", "memcpy/fill/misc", "other"]

wb = Workbook()
hdr_fill = PatternFill("solid", fgColor="D9E1F2")
nvfp4_fill = PatternFill("solid", fgColor="FCE4D6")
bold = Font(bold=True)

def style_header(ws, row=1):
    for c in ws[row]:
        if c.value: c.font = bold; c.fill = hdr_fill

# ---- Sheet 1: overview ----
ws = wb.active; ws.title = "对比总览"
ws.append(["场景", "batch", "GPU busy (ms)", "wall (ms)", "GPU 占用率", "总 token 数", "每步 GPU busy (ms)", "每 token GPU busy (us)", "吞吐 tok/s (wall)", "吞吐 tok/s (busy)", "NVFP4 时间占比"])
for (sc,b),r in data.items():
    nv = r["cat"].get("MoE NVFP4 (fused SiLU)",[0])[0]
    ws.append([sc, b, round(r["busy_us"]/1e3,2), r["wall_ms"], f"{100*r['busy_us']/1e3/r['wall_ms']:.0f}%",
               r["toks"], round(r["busy_us"]/1e3/r["steps"],2), round(r["busy_us"]/r["toks"],1),
               round(r["toks"]/r["wall_ms"]*1e3), round(r["toks"]/r["busy_us"]*1e6),
               f"{100*nv/r['busy_us']:.1f}%"])
for row in ws.iter_rows(min_row=2):
    row[10].fill = nvfp4_fill
style_header(ws)

# ---- Sheets 2/3: breakdown ----
for sc in ["prefill", "decode"]:
    ws = wb.create_sheet(f"{sc} breakdown")
    ws.append(["模块"] + [f"b{b} 时间(ms)", f"b{b} 占比", f"b{b} 每token(us)", f"b{b} 每层(us)"]*0 + [])
    ws.delete_rows(1)
    header = ["模块", "b1 时间(ms)", "b1 占比", "b1 每token(us)", "b1 每层(us)",
              "b64 时间(ms)", "b64 占比", "b64 每token(us)", "b64 每层(us)", "b64/b1 时间倍率", "b64/b1 每token效率"]
    ws.append(header); style_header(ws)
    for c in cats:
        row = [c]; vals=[]
        for b in [1,64]:
            r = data[(sc,b)]
            d, cnt = r["cat"].get(c, [0.0,0])
            layers = 2
            vals.append((d, r))
            row += [round(d/1e3,3), f"{100*d/r['busy_us']:.1f}%" if d else "0%",
                    round(d/r["toks"],2), round(d/r["steps"]/layers,1)]
        r1, r2 = vals[0][1], vals[1][1]
        d1, d2 = vals[0][0], vals[1][0]
        row += [round(d2/d1,2) if d1 else "-", round((d1/r1["toks"])/(d2/r2["toks"]),2) if d1 and d2 else "-"]
        ws.append(row)
        if c == "MoE NVFP4 (fused SiLU)":
            for cell in ws[ws.max_row]: cell.fill = nvfp4_fill
    # total row
    ws.append(["TOTAL"] + [round(data[(sc,b)]["busy_us"]/1e3,2) for b in [1,64]][:1] + ["100%",
              round(data[(sc,1)]["busy_us"]/data[(sc,1)]["toks"],1), round(data[(sc,1)]["busy_us"]/data[(sc,1)]["steps"]/2,1),
              round(data[(sc,64)]["busy_us"]/1e3,2), "100%",
              round(data[(sc,64)]["busy_us"]/data[(sc,64)]["toks"],1), round(data[(sc,64)]["busy_us"]/data[(sc,64)]["steps"]/2,1), "-", "-"])
    for cell in ws[ws.max_row]: cell.font = bold

# ---- Sheet 4: NVFP4 analysis ----
ws = wb.create_sheet("NVFP4 分析")
rows = [
 ["指标", "prefill b1", "prefill b64", "decode b1", "decode b64"],
 ["NVFP4 kernel 时间 (ms)", 1.973, 20.610, 3.666, 84.197],
 ["NVFP4 kernel 变体", "MoEDynamicKernelSilu", "MoEDynamicKernelSilu", "MoEMicroKernelSilu", "MoEDynamicKernelSilu (batch>=阈值切换)"],
 ["NVFP4 时间占比 (GPU busy)", "56.2%", "63.8%", "13.4%", "63.1%"],
 ["NVFP4 每层每 token (us)", 61.6, 10.1, 114.6, 41.1],
 ["覆盖范围", "MoE routed experts w1/w2/w3 GEMM + SiLU 融合 (唯一 NVFP4 算子)", "", "", ""],
 ["模型字节占比", "88.8% (145.1/163.5 GiB)", "", "", ""],
 ["FLOPs 占比 (43层解析)", "~51%", "", "", ""],
 ["decode 每token权重读 (NVFP4 vs FP8 vs BF16)", "57 vs 100 vs 201 MB/层", "", "", ""],
]
for r in rows: ws.append(r)
style_header(ws)

for ws in wb.worksheets:
    for i, col in enumerate(ws.columns, 1):
        w = max((len(str(c.value)) for c in col if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(i)].width = min(w+2, 60)

out = "/home/tina/DEV/microbench/results/DeepSeekV4_NVFP4_profiling_comparison.xlsx"
wb.save(out)
print("saved", out)

# also print text summary
for sc in ["prefill","decode"]:
    print(f"\n== {sc} ==")
    for b in [1,64]:
        r=data[(sc,b)]
        nv=r["cat"].get("MoE NVFP4 (fused SiLU)",[0])[0]
        print(f" b{b}: busy {r['busy_us']/1e3:.2f}ms wall {r['wall_ms']}ms nvfp4 {100*nv/r['busy_us']:.1f}% per_tok {r['busy_us']/r['toks']:.1f}us")
